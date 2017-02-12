# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html


from scrapy import signals
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, colors, PatternFill, Side, Border


class MarathonPipeline(object):
    def process_item(self, item, spider):
        return item


class ExcelPipeline(object):

    def __init__(self):
        self.col_chars = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        self.titles = [
            '名称',
            '日期',
            '比赛项目',
            '比赛规模',
            '比赛时间',
            '比赛城市',
            '路线类型',
            '赛事官网',
            '联系电话',
            '微信公众号',
            '主办单位',
            '承办单位',
            '推广单位',
            '平均温度',
            '平均湿度',
        ]
        self.wb = None
        self.ws = None
        # self.font_main_title = Font(name='Arial', size=28, color=colors.BLACK, bold=True)
        self.font_title = Font(name='Arial', size=10, color=colors.BLACK, bold=True)
        self.font_content = Font(name='Arial', size=10, color=colors.BLACK, bold=False)
        self.alignment_title = Alignment(horizontal='center', vertical='center')
        self.alignment_content = Alignment(horizontal='left', vertical='center')
        # self.fill_mail_title = PatternFill(bgColor='64B5F6')
        self.fill_odd = PatternFill(patternType='solid', fgColor='E8F5E9')
        self.fill_even = PatternFill(patternType='solid', fgColor='C8E6C9')

    @classmethod
    def from_crawler(cls, crawler):
        pipeline = cls()
        crawler.signals.connect(pipeline.spider_opened, signals.spider_opened)
        crawler.signals.connect(pipeline.spider_closed, signals.spider_closed)
        return pipeline

    def spider_opened(self, spider):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(self.titles)

    def spider_closed(self, spider):
        for row in range(1, self.ws.max_row+1):
            for col in range(1, self.ws.max_column+1):
                cell_index = '%s%d' % (self.gen_col_char(col), row)
                self.ws[cell_index].font = self.font_content if row is not 1 else self.font_title
                self.ws[cell_index].alignment = self.alignment_content if row is not 1 else self.alignment_title
                if row % 2 is 0:
                    # even
                    self.ws[cell_index].fill = self.fill_even
                else:
                    # odd
                    self.ws[cell_index].fill = self.fill_odd
        self.wb.save('marathon_2016.xlsx')

    def process_item(self, item, spider):
        line = [
            item['name'] if 'name' in item else '/',
            item['date'] if 'date' in item else '/',
            item['xiang_mu'] if 'xiang_mu' in item else '/',
            item['gui_mo'] if 'gui_mo' in item else '/',
            item['shi_jian'] if 'shi_jian' in item else '/',
            item['cheng_shi'] if 'cheng_shi' in item else '/',
            item['lei_xing'] if 'lei_xing' in item else '/',
            item['guan_wang'] if 'guan_wang' in item else '/',
            item['dian_hua'] if 'dian_hua' in item else '/',
            item['wechat_number'] if 'wechat_number' in item else '/',
            item['zhu_ban'] if 'zhu_ban' in item else '/',
            item['cheng_ban'] if 'cheng_ban' in item else '/',
            item['tui_guang'] if 'tui_guang' in item else '/',
            item['wen_du'] if 'wen_du' in item else '/',
            item['shi_du'] if 'shi_du' in item else '/',
        ]
        l = self.ws.append(line)
        return item

    def gen_col_char(self, index):
        if index is 0:
            return ''
        return self.gen_col_char(index / 16) + (self.col_chars[index % 16 - 1])
